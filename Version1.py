from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side
import os

archivo_registro = 'registro.xlsx'
archivo_compras = 'REGISTRO DE COMPRAS 2024.xlsx'
meses = {
    2: 'FEB',
    3: 'MAR',
    4: 'ABR',
    5: 'MAY',
    6: 'JUN',
    7: 'JUL',
    8: 'AGO',
    9: 'SEP'
}

# Definir el estilo de borde negro
border = Border(
    left=Side(border_style='thin', color='000000'),
    right=Side(border_style='thin', color='000000'),
    top=Side(border_style='thin', color='000000'),
    bottom=Side(border_style='thin', color='000000')
)

# Función para inicializar registro.xlsx
def inicializar_registro():
    if not os.path.exists(archivo_registro):
        wb = Workbook()
        ws = wb.active
        ws.append(["Razon Social", "RUT", "Direccion", "Tipo Comprobante", "Fecha", "Codigo Cuenta", "Detalle", "Monto"])
        wb.save(archivo_registro)

# Función para inicializar REGISTRO DE COMPRAS 2024.xlsx
def inicializar_compras():
    if not os.path.exists(archivo_compras):
        wb = Workbook()
        sheet_names = ['FEB', 'MAR', 'ABR', 'MAY', 'JUN', 'JUL', 'AGO', 'SEP']
        for idx, sheet in enumerate(sheet_names):
            if idx == 0:
                ws = wb.active
                ws.title = sheet
            else:
                ws = wb.create_sheet(title=sheet)
            ws.append([
                "N°", "Tipo Doc", "Tipo Compra", "RUT Proveedor", "Razon Social",
                "Folio", "Fecha Docto", "Monto Exento", "Monto Neto",
                "Monto IVA Recuperable", "Monto Total"
            ])
            # Aplicar borde a la fila de encabezados
            for col in range(1, 12):
                ws.cell(row=1, column=col).border = border
        wb.save(archivo_compras)

# Función para obtener el próximo número en una hoja
def obtener_proximo_numero(ws):
    if ws.max_row < 2:
        return 1
    else:
        ultimo_num = ws.cell(row=ws.max_row, column=1).value
        return ultimo_num + 1 if isinstance(ultimo_num, int) else 1

# Inicializar archivos si no existen
inicializar_registro()
inicializar_compras()

while True:
    razon_social = input("Ingrese la razon social: ")
    print("La razon social es:", razon_social)

    rut = input("Ingrese el rut: ")
    print("El rut es:", rut)

    direccion = input("Ingrese la direccion: ")
    print("La direccion es:", direccion)

#agregar tipo 3: traspaso (centralización libro de compras)
#movimientos que no necesariamente signifquen gasto o ingreso
#llena el libro diario, con el que obtenemos el balance y estado de resultado
    print("Elija el tipo de comprobante (1: Ingreso 2: Egreso)")
    tipo_comprobante = input("Ingrese el tipo de comprobante: ")
    while tipo_comprobante not in ["1", "2"]:
        print("Error, ingrese un tipo de comprobante valido")
        tipo_comprobante = input("Ingrese el tipo de comprobante: ")

    fecha_valida = False
    while not fecha_valida:
        fecha = input("Ingrese la fecha (DD/MM/YYYY): ")
        try:
            fecha_obj = datetime.strptime(fecha, "%d/%m/%Y")
            fecha_valida = True
        except ValueError:
            print("Error, ingrese una fecha valida en el formato DD/MM/YYYY")

#el codigo de cuenta corresponde a x tipo segun su numero
    print("La fecha es:", fecha)
    codigo_cuenta = input("Ingrese el codigo de cuenta: ")
    print("El codigo de cuenta es:", codigo_cuenta)


    detalle = input("Ingrese el detalle (Compra/Venta): ").capitalize()
    while detalle not in ["Compra", "Venta"]:
        print("Error, ingrese un detalle valido")
        detalle = input("Ingrese el detalle (Compra/Venta): ").capitalize()


    if tipo_comprobante == "1":
        debe = float(input("Ingrese el monto: "))
        print("El monto es:", debe)
        monto = debe
        tipo = "Debe"
    elif tipo_comprobante == "2":
        haber = float(input("Ingrese el monto: "))
        print("El monto es:", haber)
        monto = haber
        tipo = "Haber"

    try:
        # Guardar en registro.xlsx
        wb_registro = load_workbook(archivo_registro)
        ws_registro = wb_registro.active
        ws_registro.append([razon_social, rut, direccion, tipo, fecha, codigo_cuenta, detalle, monto])
        wb_registro.save(archivo_registro)
        print("Comprobante agregado exitosamente en registro.xlsx.")
    except Exception as e:
        print(f"Ocurrió un error al guardar el comprobante en registro.xlsx: {e}")

    try:
        # Determinar el mes y la hoja correspondiente
        mes = fecha_obj.month
        if mes in meses:
            hoja = meses[mes]
        else:
            print(f"Mes {mes} no está soportado para REGISTRO DE COMPRAS 2024.xlsx.")
            continuar = input("¿Desea agregar otro comprobante? (s/n): ").lower()
            if continuar != 's':
                break
            else:
                continue

        # Cargar REGISTRO DE COMPRAS 2024.xlsx
        wb_compras = load_workbook(archivo_compras)
        ws_compras = wb_compras[hoja]

        # Obtener el próximo número
        proximo_num = obtener_proximo_numero(ws_compras)

        # Calcular montos
        monto_neto = round(monto * 0.81, 2)  # Haber - 19%
        monto_iva = round(monto * 0.19, 2)   # 19% del haber
        monto_total = monto                   # Monto total es el haber

        # Agregar los datos al Libro de Compras empezando desde la columna B
        ws_compras.append([
            "",                      # Columna A vacía
            proximo_num,             # N°
            33,                      # Tipo Doc
            "Del Giro",              # Tipo Compra
            rut,                     # RUT Proveedor
            razon_social,            # Razon Social
            codigo_cuenta,           # Folio
            fecha,                   # Fecha Docto
            "",                      # Monto Exento
            monto_neto,              # Monto Neto
            monto_iva,               # Monto IVA Recuperable
            monto_total              # Monto Total
        ])

        # Obtener el número de la nueva fila
        new_row = ws_compras.max_row

        # Aplicar bordes a cada celda de la nueva fila empezando desde la columna B
        for col in range(1, 12):  # Columnas 1 a 11
            cell = ws_compras.cell(row=new_row, column=col + 1)
            cell.border = border

        # Guardar REGISTRO DE COMPRAS 2024.xlsx
        wb_compras.save(archivo_compras)
        print(f"Comprobante agregado exitosamente en REGISTRO DE COMPRAS 2024.xlsx en la hoja {hoja}.")
    except PermissionError:
        print("Error: No se pudo acceder a 'REGISTRO DE COMPRAS 2024.xlsx'. Asegúrate de que el archivo no esté abierto en otra aplicación.")
    except Exception as e:
        print(f"Ocurrió un error al guardar el comprobante en REGISTRO DE COMPRAS 2024.xlsx: {e}")

    continuar = input("¿Desea agregar otro comprobante? (s/n): ").lower()
    if continuar != 's':
        break
    
    
#Las cuentas que comienzan por 1 corresponden a activos
#Las cuentas que comienzan por 2 pasivos
#Las cuentas 3 a gastos
#las cuentas 4 a ingresos

#Pendiente revisar la alimentación de libro diario
#Contiene fecha,codigo de cuenta, la cuenta, glosa, debe y haber
#agrupar todos los haberes con ellos y todos los debe con ellos hacia arriba

#a fin de mes, se bloquea la edicion del libro y se inicia uno nuevo
