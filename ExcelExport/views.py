import openpyxl
from django.shortcuts import render, redirect
from django.http import HttpResponse
from .forms import VentasForm, ComprasForm, LibroDiarioForm
import pandas as pd
from django.http import FileResponse
import os
from django.conf import settings
from decimal import Decimal
from datetime import datetime
from io import BytesIO
from itertools import zip_longest  


def descargar_libro(request, tipo):
    if tipo == "ventas":
        path = os.path.join(settings.BASE_DIR, "REGISTRO DE VENTAS 2024.xlsx")
        filename = "REGISTRO DE VENTAS 2024.xlsx"
    elif tipo == "compras":
        path = os.path.join(settings.BASE_DIR, "REGISTRO DE COMPRAS 2024.xlsx")
        filename = "REGISTRO DE COMPRAS 2024.xlsx"
    elif tipo == "libro_diario":
        path = os.path.join(settings.BASE_DIR, "LDE.xlsx")
        filename = "LDE.xlsx"
    elif tipo == "balance":
        # Generar el archivo Balance a partir de la hoja 'Balance' en 'Contab 2024.xlsx'
        contab_path = os.path.join(settings.BASE_DIR, "Contab 2024.xlsx")
        if not os.path.exists(contab_path):
            return HttpResponse("El archivo 'Contab 2024.xlsx' no existe.", status=404)

        wb = openpyxl.load_workbook(contab_path,data_only=True)
        if 'BALANCE' not in wb.sheetnames:
            return HttpResponse("La hoja 'BALANCE' no existe en 'BALANCE.xlsx'.", status=404)

        balance_sheet = wb['BALANCE']

        # Crear un nuevo libro de Excel y copiar la hoja 'Balance'
        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active
        new_ws.title = 'BALANCE'

        for row in balance_sheet.iter_rows(values_only=True):
            new_ws.append(row)
        output = BytesIO()
        new_wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=Balance.xlsx'
        return response
    else:
        return HttpResponse("Tipo de archivo no válido.", status=400)

    if os.path.exists(path):
        return FileResponse(open(path, 'rb'), as_attachment=True, filename=filename)
    else:
        return HttpResponse("El archivo no existe.", status=404)

def libro_diario(request):
    if request.method == "POST":
        form = LibroDiarioForm(request.POST)
        if form.is_valid():
            lde_path = os.path.join(settings.BASE_DIR, "LDE.xlsx")
            wb = openpyxl.load_workbook(lde_path)
            ws = wb.active

            # Buscar la última fila con datos reales a partir de la fila 3
            last_data_row = None
            for row in range(3, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=1).value  
                if cell_value is None:
                    last_data_row = row - 1
                    break
            if last_data_row is None:
                last_data_row = ws.max_row

            # Determinar la nueva fila para insertar datos
            new_row = last_data_row + 1

            # Obtener los valores del formulario
            fechas = request.POST.getlist('fecha')
            tipo_movimientos = request.POST.getlist('tipo_movimiento')
            nombre_cuentas = request.POST.getlist('nombre_cuenta')
            glosas = request.POST.getlist('glosa')
            debes = request.POST.getlist('debe_hidden')
            habers = request.POST.getlist('haber_hidden')

            # Insertar cada registro en el Excel utilizando zip_longest para evitar IndexError
            for fecha, tipo_movimiento, nombre_cuenta, glosa, debe, haber in zip_longest(
                fechas, tipo_movimientos, nombre_cuentas, glosas, debes, habers, fillvalue='0'
            ):
                fecha_formateada = datetime.strptime(fecha, '%Y-%m-%d').strftime('%d-%m-%Y')

                debe = Decimal(debe) if debe else Decimal('0')
                haber = Decimal(haber) if haber else Decimal('0')

                # Obtener 'Comp' inicial
                comp = tipo_movimiento[0].upper()

                # Calcular 'N°' correspondiente al 'Comp' actual
                comp_numbers = []
                for row_num in range(3, last_data_row + 1):
                    cell_comp = ws.cell(row=row_num, column=2).value  
                    cell_number = ws.cell(row=row_num, column=3).value  
                    if cell_comp == comp and isinstance(cell_number, int):
                        comp_numbers.append(cell_number)
                if comp_numbers:
                    next_number = max(comp_numbers) + 1
                else:
                    next_number = 1

                # Insertar los datos en la nueva fila
                ws.cell(row=new_row, column=1).value = fecha_formateada
                ws.cell(row=new_row, column=2).value = comp
                ws.cell(row=new_row, column=3).value = next_number
                ws.cell(row=new_row, column=4).value = nombre_cuenta
                ws.cell(row=new_row, column=5).value = glosa
                ws.cell(row=new_row, column=6).value = debe
                ws.cell(row=new_row, column=7).value = haber

                new_row += 1  # Incrementar para la siguiente fila

            # Actualizar las fórmulas de suma en F1 y G1
            ws['F1'].value = f"=SUM(F3:F{new_row - 1})"
            ws['G1'].value = f"=SUM(G3:G{new_row - 1})"

            # Guardar el archivo Excel
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            response = HttpResponse(
                output,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=LDE.xlsx'
            return response
    else:
        form = LibroDiarioForm()
    return render(request, 'libro_diario.html', {'form': form})
def home(request):
    return render(request, 'home.html')

def ver_registro(request):
    # Lista de libros disponibles
    libros = [
        {'nombre': 'Registro de Ventas 2024', 'tipo': 'ventas'},
        {'nombre': 'Registro de Compras 2024', 'tipo': 'compras'},
        {'nombre': 'Libro Diario', 'tipo': 'libro_diario'},
        {'nombre': 'Balance', 'tipo': 'balance'},
    ]
    return render(request, 'ver_registro.html', {'libros': libros})