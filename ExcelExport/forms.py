from django import forms
import os
import openpyxl
from django.conf import settings

class VentasForm(forms.Form):
    tipo_documento = forms.CharField(label="Tipo de Documento", max_length=50)
    total_documentos = forms.IntegerField(label="Total de Documentos")
    monto_exento = forms.DecimalField(label="Monto Exento", max_digits=10, decimal_places=2, required=False)
    monto_neto = forms.DecimalField(label="Monto Neto", max_digits=10, decimal_places=2)
    monto_iva = forms.DecimalField(label="Monto IVA", max_digits=10, decimal_places=2)
    monto_total = forms.DecimalField(label="Monto Total", max_digits=10, decimal_places=2)

class ComprasForm(forms.Form):
    numero_operacion = forms.IntegerField(label="Número de Operación")
    tipo_documento = forms.CharField(label="Tipo de Documento", max_length=50)
    tipo_compra = forms.CharField(label="Tipo de Compra", max_length=50)
    rut_proveedor = forms.CharField(label="RUT Proveedor", max_length=12)
    razon_social = forms.CharField(label="Razón Social", max_length=100)
    folio = forms.IntegerField(label="Folio")
    fecha_documento = forms.DateField(label="Fecha del Documento", widget=forms.DateInput(attrs={'type': 'date'}))
    monto_exento = forms.DecimalField(label="Monto Exento", max_digits=10, decimal_places=2, required=False)
    monto_total = forms.DecimalField(label="Monto Total", max_digits=10, decimal_places=2)



class LibroDiarioForm(forms.Form):
    fecha = forms.DateField(
        label="Fecha",
        widget=forms.DateInput(attrs={'type': 'date', 'class': 'form-control'})
    )
    tipo_movimiento = forms.ChoiceField(
        label="Tipo de Movimiento",
        choices=[
            ('Traspaso', 'Traspaso'),
            ('Egreso', 'Egreso'),
            ('Ingreso', 'Ingreso'),
        ],
        widget=forms.Select(attrs={'class': 'form-control'})
    )
    nombre_cuenta = forms.ChoiceField(
        choices=[],
        widget=forms.Select(attrs={'class': 'form-control'})
    
        )  # Cambiado a ChoiceField
    glosa = forms.CharField(
        label="Glosa o Detalle",
        max_length=200,
        widget=forms.TextInput(attrs={'class': 'form-control'})
    )
    debe = forms.DecimalField(
        label="Debe",
        max_digits=15,
        decimal_places=2,
        required=False,
        widget=forms.NumberInput(attrs={'class': 'form-control'})
    )
    haber = forms.DecimalField(
        label="Haber",
        max_digits=15,
        decimal_places=2,
        required=False,
        widget=forms.NumberInput(attrs={'class': 'form-control'})
    )
    def __init__(self, *args, **kwargs):
        super(LibroDiarioForm, self).__init__(*args, **kwargs)
        import openpyxl
        import os
        from django.conf import settings

        contab_path = os.path.join(settings.BASE_DIR, "Contab 2024.xlsx")
        if os.path.exists(contab_path):
            wb = openpyxl.load_workbook(contab_path)
            if 'CUENTAS' in wb.sheetnames:
                cuentas_sheet = wb['CUENTAS']
                opciones = []
                for row in cuentas_sheet.iter_rows(min_row=6):
                    codigo = row[1].value  # Columna B (índice 1)
                    cuenta = row[4].value  # Columna E (índice 4)
                    if codigo and cuenta:
                        codigo_formateado = int(codigo)
                        opcion = f"{codigo_formateado}- {cuenta}"
                        opciones.append((cuenta, opcion))
                self.fields['nombre_cuenta'].choices = opciones
            else:
                print("La hoja 'CUENTAS' no existe en el libro.")
                self.fields['nombre_cuenta'].choices = []
        else:
            print("El archivo 'Contab 2024.xlsx' no existe.")
            self.fields['nombre_cuenta'].choices = []